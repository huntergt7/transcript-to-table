import io
import re
import html
from typing import List, Dict, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Border, Side

# =========================
# Toggleable debug flag
# =========================
DEBUG_ON = False  # Set True to default-enable parser trace and internal logging


# =========================
# Parsing helpers & regexes
# =========================

WRAP_CHARS = " \t\r\n\"'{}<>“”‘’"

def _strip_wrappers(s: str) -> str:
    return s.strip(WRAP_CHARS)

def _clean_quote(s: str) -> str:
    s = s.strip()
    s = re.sub(r"^[\-–—:·•*#> ]+", "", s).strip()  # strip leading punctuation/dashes
    s = _strip_wrappers(s)
    s = re.sub(r"\s+", " ", s)
    return s

def _is_meaningful(text: str) -> bool:
    return bool(text and re.search(r"\w", text))

# Timestamp regexes
_TS_TOKEN_RE = re.compile(r"^(\d{1,2}):(\d{2})(?:\.\d+)?$")
_TS_RE_ANYWHERE = re.compile(r"\b(\d{1,2}):(\d{2})(?::(\d{2}))?(?:\.\d+)?\b")

def _to_mmss(mm: int, ss: int, shift_seconds: float = 0.0) -> str:
    total = mm * 60 + ss
    total = int(max(0, total - float(shift_seconds)))
    m, s = divmod(total, 60)
    return f"{m:02d}:{s:02d}"

# Robust detection of WEBVTT cue arrows: "-->", "–>", "—>"
CUE_ARROW_RE = re.compile(r"(?:--|–|—)\>", re.UNICODE)

# Name normalization & replacement (emoji/punctuation-safe)
def _normalize_name_input(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r"\s+", " ", name)
    return name

# Fix common mojibake (e.g., copy/paste "â€¦" → "…")
_MOJIBAKE_FIXES = {
    "â€¦": "…", "â€”": "—", "â€“": "–", "Â ": " ",
}
def _normalize_mojibake(s: str) -> str:
    for bad, good in _MOJIBAKE_FIXES.items():
        s = s.replace(bad, good)
    return s

# boundary characters (no need for \b; supports emojis/punct)
_BOUNDARY = r"[^\S\r\n]|[\-–—:;,/\\\[\]\(\)\{\}“”\"'<>…!?.]"

def _build_name_inner_pattern(name: str) -> Optional[str]:
    if not name:
        return None
    tokens = _normalize_name_input(name).split()
    return r"\s+".join(map(re.escape, tokens))

def _whole_name_captor(name: str) -> Optional[re.Pattern]:
    inner = _build_name_inner_pattern(name)
    if not inner:
        return None
    pattern = rf"(^|{_BOUNDARY})({inner})(?=({_BOUNDARY}|$))"
    return re.compile(pattern, re.IGNORECASE | re.UNICODE)

def _safe_replace_whole_name(text: str, name: str, replacement: str) -> str:
    pat = _whole_name_captor(name)
    if not pat:
        return text
    def _sub(m: re.Match) -> str:
        prefix = m.group(1) or ""
        return f"{prefix}{replacement}"
    return pat.sub(_sub, text)

# Speaker patterns (emoji/punctuation-capable)
_SPEAKER_BRACKET_RE    = re.compile(r"^\s*\[(?P<name>[^\]]+)\]\s*(?P<rest>.*)$")
_SPEAKER_PAREN_ANY_RE  = re.compile(r"^\s*(?P<name>.+?)\s*\((?P<ts>[^)]*?)\)\s*(?P<rest>.*)$")
_SPEAKER_DELIM_ANY_RE  = re.compile(r"^\s*(?P<name>.+?)\s*[:：\-–—]\s+(?P<rest>.*)$")

def _build_bare_token_patterns(counselor_name: str, client_name: str):
    pats = []
    def add(name: str, label: str):
        name = _normalize_name_input(name)
        if not name:
            return
        inner = _build_name_inner_pattern(name)
        pattern = rf"^\s*({inner})(?=({_BOUNDARY}|$))\s*(?P<rest>.*)$"
        pats.append((label, re.compile(pattern, re.IGNORECASE | re.UNICODE)))
    # user-provided names
    add(counselor_name, "Couns")
    add(client_name, "Client")
    # already-anonymized tokens
    add("Couns", "Couns")
    add("Client", "Client")
    return pats


# =========================
# Core parser
# =========================

def parse_dialogue_text(
    text: str,
    counselor_name: str,
    client_name: str,
    shift_seconds: float = 0.0,
    trace: bool = False
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Parse free-form transcript text into a DataFrame with:
      Timestamp | Speaker | Quote | Tag

    Auto-detects:
      - WEBVTT cue lines "HH:MM:SS.mmm --> HH:MM:SS.mmm" (first timestamp carried to next quote)
      - Timestamp-only lines above quotes
      - Name (timestamp), [Name], Name: / Name —, and bare tokens ("Client ...", "Couns ...")
    """
    logs: List[str] = []
    def log(msg):
        if trace:
            logs.append(msg)

    counselor_name = _normalize_name_input(counselor_name)
    client_name    = _normalize_name_input(client_name)

    lines = text.splitlines(keepends=False)
    rows: List[Dict[str, str]] = []
    pending_speaker: Optional[str] = None
    pending_ts: Optional[str] = None

    bare_pats = _build_bare_token_patterns(counselor_name, client_name)

    def map_public(name: Optional[str]) -> Optional[str]:
        if not name:
            return None
        nm = _normalize_name_input(name).casefold()
        if nm in (counselor_name.casefold(), "couns"):
            return "Couns"
        if nm in (client_name.casefold(), "client"):
            return "Client"
        return name

    def _append_or_add(speaker_public: str, q: str, ts_for_row: str):
        nonlocal rows
        if rows and rows[-1]["Speaker"] == speaker_public:
            prev_len = len(rows[-1]["Quote"])
            rows[-1]["Quote"] = (rows[-1]["Quote"] + " " + q).strip()
            if speaker_public == "Client":
                rows[-1]["Timestamp"] = ""  # enforce blank for Client
            log(f"    ↳ APPEND to {speaker_public} | +{len(rows[-1]['Quote']) - prev_len} chars")
        else:
            rows.append({"Timestamp": ts_for_row, "Speaker": speaker_public, "Quote": q})
            log(f"    ↳ NEW ROW: {speaker_public} | TS={ts_for_row} | +{len(q)} chars")

    for idx, raw in enumerate(lines, start=1):
        # normalize encoding artifacts and HTML entities
        line = _normalize_mojibake(html.unescape(raw)).replace("\ufeff", "").strip()
        if not line:
            continue

        # Skip WEBVTT header
        if line.upper().startswith("WEBVTT"):
            log(f"[L{idx}] Skip WEBVTT header")
            continue

        # --- A) WEBVTT cue lines (e.g., "00:52:14.000 --> 00:52:16.000") ---
        if CUE_ARROW_RE.search(line):
            anym = _TS_RE_ANYWHERE.search(line)  # take the FIRST timestamp only
            if anym:
                if anym.group(3):  # HH:MM:SS(.ms)
                    mm, ss = int(anym.group(2)), int(anym.group(3))
                else:              # MM:SS(.ms)
                    mm, ss = int(anym.group(1)), int(anym.group(2))
                pending_ts = _to_mmss(mm, ss, shift_seconds)
                log(f"[L{idx}] VTT cue → carry TS={pending_ts}")
            continue

        # --- B) Timestamp-only line (no other meaningful text) ---
        ts_probe = _TS_RE_ANYWHERE.search(line)
        if ts_probe:
            s, e = ts_probe.span()
            leftover = (line[:s] + " " + line[e:]).strip()
            leftover_clean = re.sub(r"\s+", " ", _strip_wrappers(re.sub(r"^[\-\–—:·•*#> ]+", "", leftover))).strip()
            if not _is_meaningful(leftover_clean):
                if ts_probe.group(3):
                    mm, ss = int(ts_probe.group(2)), int(ts_probe.group(3))
                else:
                    mm, ss = int(ts_probe.group(1)), int(ts_probe.group(2))
                pending_ts = _to_mmss(mm, ss, shift_seconds)
                log(f"[L{idx}] TS-only line → carry TS={pending_ts}")
                continue

        # --- C) Name (timestamp) at start ---
        m = _SPEAKER_PAREN_ANY_RE.match(line)
        if m:
            raw_name, ts_token, rest = m.group("name"), m.group("ts"), m.group("rest")
            anym = _TS_RE_ANYWHERE.search(ts_token)
            if anym:
                if anym.group(3):
                    mm, ss = int(anym.group(2)), int(anym.group(3))
                else:
                    mm, ss = int(anym.group(1)), int(anym.group(2))
                pending_speaker = _normalize_name_input(raw_name)
                pending_ts = _to_mmss(mm, ss, shift_seconds)
                log(f"[L{idx}] Name(ts) → pending speaker='{pending_speaker}', TS={pending_ts}")

                q = _clean_quote(rest)
                if _is_meaningful(q):
                    speaker_public = map_public(pending_speaker) or ""
                    q = _safe_replace_whole_name(q, counselor_name, "Couns")
                    q = _safe_replace_whole_name(q, client_name, "Client")
                    ts_for_row = "" if speaker_public == "Client" else (pending_ts or "")
                    _append_or_add(speaker_public, q, ts_for_row)
                    pending_speaker = None
                    pending_ts = None
                continue

        # --- D) [Name] ... ---
        m = _SPEAKER_BRACKET_RE.match(line)
        if m:
            raw_name, rest = m.group("name"), m.group("rest")
            pending_speaker = _normalize_name_input(raw_name)
            anym = _TS_RE_ANYWHERE.search(rest)
            if anym:
                if anym.group(3):
                    mm, ss = int(anym.group(2)), int(anym.group(3))
                else:
                    mm, ss = int(anym.group(1)), int(anym.group(2))
                pending_ts = _to_mmss(mm, ss, shift_seconds)
                s, e = anym.span()
                rest = (rest[:s] + " " + rest[e:]).strip()
                log(f"[L{idx}] [Name] with inline ts → pending TS={pending_ts}")

            q = _clean_quote(rest)
            if _is_meaningful(q):
                speaker_public = map_public(pending_speaker) or ""
                q = _safe_replace_whole_name(q, counselor_name, "Couns")
                q = _safe_replace_whole_name(q, client_name, "Client")
                ts_for_row = "" if speaker_public == "Client" else (pending_ts or "")
                _append_or_add(speaker_public, q, ts_for_row)
                pending_speaker = None
                pending_ts = None
            continue

        # --- E) Name :/—/- rest ---
        m = _SPEAKER_DELIM_ANY_RE.match(line)
        if m:
            raw_name, rest = m.group("name"), m.group("rest")
            pending_speaker = _normalize_name_input(raw_name)
            anym = _TS_RE_ANYWHERE.search(rest)
            if anym:
                if anym.group(3):
                    mm, ss = int(anym.group(2)), int(anym.group(3))
                else:
                    mm, ss = int(anym.group(1)), int(anym.group(2))
                pending_ts = _to_mmss(mm, ss, shift_seconds)
                s, e = anym.span()
                rest = (rest[:s] + " " + rest[e:]).strip()
                log(f"[L{idx}] Name: with inline ts → pending TS={pending_ts}")

            q = _clean_quote(rest)
            if _is_meaningful(q):
                speaker_public = map_public(pending_speaker) or ""
                q = _safe_replace_whole_name(q, counselor_name, "Couns")
                q = _safe_replace_whole_name(q, client_name, "Client")
                ts_for_row = "" if speaker_public == "Client" else (pending_ts or "")
                _append_or_add(speaker_public, q, ts_for_row)
                pending_speaker = None
                pending_ts = None
            continue

        # --- F) bare tokens (always on): "Client …", "Couns …", or exact names w/ emojis ---
        matched_bare = False
        for label, pat in bare_pats:
            bm = pat.match(line)
            if bm:
                rest = bm.group("rest")
                pending_speaker = label
                anym = _TS_RE_ANYWHERE.search(rest)
                if anym:
                    if anym.group(3):
                        mm, ss = int(anym.group(2)), int(anym.group(3))
                    else:
                        mm, ss = int(anym.group(1)), int(anym.group(2))
                    pending_ts = _to_mmss(mm, ss, shift_seconds)
                    s, e = anym.span()
                    rest = (rest[:s] + " " + rest[e:]).strip()
                    log(f"[L{idx}] Bare '{label}' with inline ts → pending TS={pending_ts}")
                q = _clean_quote(rest)
                if _is_meaningful(q):
                    speaker_public = map_public(pending_speaker) or ""
                    q = _safe_replace_whole_name(q, counselor_name, "Couns")
                    q = _safe_replace_whole_name(q, client_name, "Client")
                    ts_for_row = "" if speaker_public == "Client" else (pending_ts or "")
                    _append_or_add(speaker_public, q, ts_for_row)
                    pending_speaker = None
                    pending_ts = None
                matched_bare = True
                break
        if matched_bare:
            continue

        # --- G) continuation ---
        q = _clean_quote(line)
        if not _is_meaningful(q):
            continue
        if rows and (pending_speaker is None):
            prev_len = len(rows[-1]["Quote"])
            rows[-1]["Quote"] = (rows[-1]["Quote"] + " " + q).strip()
            if rows[-1]["Speaker"] == "Client":
                rows[-1]["Timestamp"] = ""
            log(f"[L{idx}] APPEND to {rows[-1]['Speaker']} | +{len(rows[-1]['Quote']) - prev_len} chars")
        else:
            speaker_public = map_public(pending_speaker) or ""
            q = _safe_replace_whole_name(q, counselor_name, "Couns")
            q = _safe_replace_whole_name(q, client_name, "Client")
            ts_for_row = "" if speaker_public == "Client" else (pending_ts or "")
            _append_or_add(speaker_public, q, ts_for_row)
            pending_speaker = None
            pending_ts = None

    # Final consolidation: merge adjacent same-speaker rows
    def collapse_consecutive_rows(rows_in: List[Dict[str, str]]) -> List[Dict[str, str]]:
        merged: List[Dict[str, str]] = []
        for r in rows_in:
            if merged and merged[-1]["Speaker"] == r["Speaker"]:
                if merged[-1]["Speaker"] == "Client":
                    merged[-1]["Timestamp"] = ""
                merged[-1]["Quote"] = (merged[-1]["Quote"] + " " + r["Quote"]).strip()
            else:
                merged.append(r.copy())
        return merged

    rows = collapse_consecutive_rows(rows)

    # Tag & enforce Client timestamps
    def _wc(s: str) -> int:
        s = s.strip()
        return 0 if not s else len(re.split(r"\s+", s))

    for r in rows:
        if r["Speaker"] == "Client":
            r["Timestamp"] = ""  # ensure empty TS for all Client rows
        r["Tag"] = "ME" if (r["Speaker"] == "Couns" and _wc(r["Quote"]) <= 3) else ""

    df = pd.DataFrame(rows, columns=["Timestamp", "Speaker", "Quote", "Tag"])
    return df, logs


# =========================
# Streamlit UI
# =========================

st.set_page_config(page_title="Transcript → Counseling Table", page_icon="📝")
st.title("📝 Video Recording Transcript Converter for COUN 633")
st.caption("Disclaimer: Use at your own risk. By interacting with this tool, you agree to the Terms of Service of Streamlit.io")
st.caption("This tool was developed with the help of Microsoft Copilot. For troubleshooting, please contact Hunter T.")
st.caption("_Last updated February 11, 2026._")

with st.sidebar:
    st.header("Settings")
    counselor_name_input = st.text_input("Counselor name (as it appears in transcript)", placeholder="<Couns Name>")
    client_name_input    = st.text_input("Client name (as it appears in transcript)",    placeholder="<Client Name>")

    counselor_name = _normalize_name_input(counselor_name_input)
    client_name    = _normalize_name_input(client_name_input)

    # Default shift set to 1.0s (your preference)
    shift_seconds = st.number_input(
        "Seconds to subtract (optional)", min_value=0, value=0, step=1,
        help="This is only necessary if you trimmed your video and need to modify the timestamps to match."
    )

    # Debug checkbox defaults from DEBUG_ON
    show_trace = False
    if DEBUG_ON:
        show_trace = st.checkbox("Show parser trace (debug)", value=DEBUG_ON)

    st.markdown("---")
    input_method = st.radio("Input method", ["Upload .txt", "Paste text"], index=0)

uploaded_text = None
uploaded_file = None

if input_method == "Upload .txt":
    uploaded_file = st.file_uploader("Upload transcript (.txt)", type=["txt"])
    if uploaded_file is not None:
        uploaded_text = uploaded_file.read().decode("utf-8", errors="ignore")
else:
    uploaded_text = st.text_area("Paste transcript text here", height=260, placeholder="Paste your transcript…")

if st.button("Parse & Generate"):
    if not uploaded_text:
        st.error("Please upload a .txt file or paste the transcript text.")
    elif not counselor_name or not client_name:
        st.error("Please enter both Counselor and Client names in the sidebar.")
    else:
        df, logs = parse_dialogue_text(
            uploaded_text,
            counselor_name=counselor_name,
            client_name=client_name,
            shift_seconds=shift_seconds,
            trace=show_trace
        )

        if df.empty:
            st.warning("No dialogue rows were parsed. Check your input and name settings.")
        else:
            st.subheader("Preview")
            st.dataframe(df, use_container_width=True, height=360)

            # Write to XLSX (in-memory) with blue font for Client rows
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, header=False, sheet_name="Dialogue")
                ws = writer.sheets["Dialogue"]
                
                # Define styles
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                base_font  = Font(name="Times New Roman", size=12)
                blue_font  = Font(name="Times New Roman", size=12, color="0000FF")  # for Client rows    
                black_side = Side(style="thin", color="000000")
                black_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)

                # Apply to all cells
                for row in ws.iter_rows(min_row=0, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.fill = white_fill     # white background
                        cell.font = base_font      # Times New Roman 12pt
                        cell.border = black_border
                
                # Then re-apply blue font to Client rows (column 2 = Speaker)
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=2).value == "Client":
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=r, column=c).font = blue_font

            output.seek(0)

            suggested_name = "converted_transcript.xlsx"
            if input_method == "Upload .txt" and uploaded_file is not None:
                base = uploaded_file.name.rsplit(".", 1)[0]
                suggested_name = f"{base}.xlsx"

            st.download_button(
                label="⬇️ Download XLSX",
                data=output.getvalue(),
                file_name=suggested_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # # Also offer a clean CSV export
            # st.download_button(
            #     label="⬇️ Download CSV",
            #     data=df.to_csv(index=False).encode("utf-8"),
            #     file_name=suggested_name.replace(".xlsx", ".csv"),
            #     mime="text/csv",
            # )

            # Quick stats
            total = len(df)
            n_client = int((df["Speaker"] == "Client").sum())
            n_couns  = int((df["Speaker"] == "Couns").sum())
            st.info(f"Rows: {total} • Client rows: {n_client} • Couns rows: {n_couns} • Tags (ME): {int((df['Tag']=='ME').sum())}")

        if show_trace and logs:
            st.subheader("Parser Trace (debug)")
            st.code("\n".join(logs)[:100000], language="text")
